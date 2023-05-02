# coding=utf-8

import openpyxl

def evaluate(param1, param2):
    flag = True
    index_rows = []
    for item in range(len(param1)):
        if param1[item].value != param2[item].value:
            flag = False
        else:
            index_rows.append(item + 1)
    return flag, index_rows

# Ruta de los ficheros
wb = openpyxl.load_workbook(".\Employee_Sample_Data.xlsx")
wb2 = openpyxl.load_workbook(".\Employee_Sample_Data_2.xlsx")
# ws = wb["Data"]           Si se tiene varias hojas, acá se pone el nombre de la hoja especifica
# ws2 = wb["Data"]           Si se tiene varias hojas, acá se pone el nombre de la hoja especifica

ws = wb.active                  # Selecciona la hoja activa
ws2 = wb2.active                # Selecciona la hoja activa

columna = ws['B']               # Accediendo a la columna
columna2 = ws2['B']             # Accediendo a la columna

if len(columna) > len(columna2):
    flag, index_rows = evaluate(columna, columna2)
else:
    flag, index_rows = evaluate(columna2, columna)

if flag:
    print("Ambas columnas son iguales.")

if index_rows:
    print("Las columnas son iguales en estas filas: ")
    for item in index_rows:
        print(item, " - ",columna[item - 1].value)

    print("\t\tHay ", (len(index_rows) - 1), " filas iguales")
else:
    print("No hay nada igual")