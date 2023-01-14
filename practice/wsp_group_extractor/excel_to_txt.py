# Guillermo Cala

# este script esta diseñado para obtener de un archivo excel
# la primera columna y convertir los numeros crudos
# a hipervinculos de whatsapp para enviar mensajes
# dejando en un txt la lista con la cantidad de contactos
# por grupo(hoja de excel)

from openpyxl import load_workbook

def onlyNumbers(number):
    res = ""
    for item in number:
        if item.isnumeric():
            res += item
    return res

links = open("prueba.txt", "w")

wb = load_workbook(".\grupos.xlsx")

for actual_ws in wb:
    cantidad = str(actual_ws.max_row - 2)
    links.write("\t\t\t" + actual_ws.title + " - Cantidad: " + cantidad + "\n") 
    ws = actual_ws

    for row in ws.iter_rows(min_row=3, max_col=1):
        for cell in row:
            temp = "https://wa.me/" + onlyNumbers(cell.value) + "\n"
            links.write(temp)
